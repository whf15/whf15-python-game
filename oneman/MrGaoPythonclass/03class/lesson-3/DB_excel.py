from openpyxl import load_workbook, Workbook

#创建新excel文档
# wb = Workbook()
# ws = wb.create_sheet('myDATA')
# # ws = wb.active
#
# ws.append(['name','年龄'])
#
# wb.save('111.xlsx')


#开启excel写入处理后的数据

result_list = [
    ['正正','28'],
    ['培培', '35'],
    ['宁夫', '29'],
]
wb = load_workbook('111.xlsx')
ws = wb['myDATA']

# for row in result_list:
#     ws.append(row)
# wb.save(filename='111.xlsx')

for x in ws.iter_rows(min_row=2):
    print(x)
    temp_row = []
    for y in x:
        temp_row.append(y.value)
    print(temp_row)
